#!/usr/bin/env python3
import os
import re
import sys
import threading
import traceback
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(
        "Missing Dependency",
        "openpyxl is required.\n\nInstall it with:\n\npip install openpyxl"
    )
    sys.exit(1)

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

# ============================================================
# PATHS
# ============================================================
BASE_DIR      = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR    = os.path.join(BASE_DIR, "source")
TEMPLATE_DIR  = os.path.join(BASE_DIR, "template")
DEST_DIR      = os.path.join(BASE_DIR, "destination")
LOG_FILE      = os.path.join(BASE_DIR, "data_extraction_log.txt")

# ============================================================
# HELPERS  (same logic as original CLI)
# ============================================================
TITLE_RE = re.compile(
    r"^(mr\.?|ms\.?|mrs\.?|miss\.?|dr\.?|prof\.?)\s+",
    re.IGNORECASE,
)


def normalize_name(name):
    if not name:
        return ""
    n = str(name).strip()
    n = TITLE_RE.sub("", n)
    n = re.sub(r"\s+", " ", n)
    return n.upper().strip()


def clean_display_name(name):
    if not name:
        return ""
    n = str(name).strip()
    n = TITLE_RE.sub("", n)
    return re.sub(r"\s+", " ", n).strip()


def is_empty(val):
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in ("", "none", "nan", "--", "null")


def combine_address(addr1, addr2):
    a1 = str(addr1).strip() if not is_empty(addr1) else ""
    a2 = str(addr2).strip() if not is_empty(addr2) else ""
    if a1 and a2:
        return f"{a1}, {a2}"
    return a1 or a2 or ""


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    s = str(val).strip()
    if is_empty(s):
        return None
    formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
        "%d/%b/%Y", "%d-%b-%Y", "%d/%B/%Y", "%d-%B-%Y",
        "%d-%b-%y", "%d/%b/%y", "%d-%m-%y", "%d/%m/%y",
        "%b/%d/%Y", "%B/%d/%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return s


def format_date(d):
    if d is None:
        return None
    if isinstance(d, date):
        return d.strftime("%d/%m/%Y")
    return str(d)


def find_header_row(ws, must_contain, search_rows=15):
    must = [m.lower() for m in must_contain]
    for r in range(1, min(search_rows, ws.max_row) + 1):
        values = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                values.append(str(v).strip().lower())
        if all(any(m in v for v in values) for m in must):
            return r
    return None


def build_header_map(ws, header_row):
    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(header_row, col).value
        if h:
            headers[str(h).strip()] = col
    return headers


def get_col(headers, *candidates):
    lower_map = {k.lower(): v for k, v in headers.items()}
    for cand in candidates:
        cand_l = cand.lower()
        if cand_l in lower_map:
            return lower_map[cand_l]
        for k, v in lower_map.items():
            if cand_l in k or k in cand_l:
                return v
    return None


def list_xlsx(folder):
    if not os.path.isdir(folder):
        return []
    return sorted(
        os.path.join(folder, f)
        for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
    )


def classify_source_files(paths):
    booking, personal = [], []
    for path in paths:
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            text = ""
            for r in range(1, min(16, ws.max_row or 1) + 1):
                for c in range(1, min(40, ws.max_column or 1) + 1):
                    v = ws.cell(r, c).value
                    if v:
                        text += " " + str(v).lower()
            wb.close()

            if "booking confirmation" in text or "enrollment no" in text:
                booking.append(path)
            elif "student personal" in text or "birth date" in text or "guardian name" in text:
                personal.append(path)
            else:
                name = os.path.basename(path).lower()
                if "booking" in name or "confirmation" in name:
                    booking.append(path)
                elif "personal" in name or "student" in name:
                    personal.append(path)
                else:
                    personal.append(path)
        except Exception:
            personal.append(path)
    return booking, personal


def load_source_one_file(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws, must_contain=["student name", "enrollment"], search_rows=20)
    if header_row is None:
        header_row = find_header_row(ws, must_contain=["student name"], search_rows=20)
    if header_row is None:
        wb.close()
        return []

    headers = build_header_map(ws, header_row)

    col_enroll  = get_col(headers, "Enrollment No.", "Enrollment No", "Enrollment Number", "Enroll No")
    col_name    = get_col(headers, "Student Name")
    col_rev_bc  = get_col(headers, "Revised/Booking Confirmation Number", "Revised Booking Confirmation Number")
    col_orig_bc = get_col(headers, "Original Booking Confirmation Number", "Booking Confirmation Number", "Booking Confirmation No")
    col_bc_date = get_col(headers, "Booking Confirmation Date")
    col_course  = get_col(headers, "Course Code")
    col_status  = get_col(headers, "Status")
    col_monthly_fee_after_discount = get_col(headers, "Monthly Fee After Discount")

    records = []
    for row in range(header_row + 1, ws.max_row + 1):
        name_raw = ws.cell(row, col_name).value if col_name else None
        if is_empty(name_raw):
            continue

        bc_num = None
        if col_rev_bc and not is_empty(ws.cell(row, col_rev_bc).value):
            bc_num = ws.cell(row, col_rev_bc).value
        elif col_orig_bc:
            bc_num = ws.cell(row, col_orig_bc).value

        records.append({
            "enrollment_no": str(ws.cell(row, col_enroll).value).strip() if col_enroll and ws.cell(row, col_enroll).value else "",
            "student_name_raw": str(name_raw).strip(),
            "student_name_norm": normalize_name(name_raw),
            "student_name_clean": clean_display_name(name_raw),
            "booking_confirmation_no": str(bc_num).strip() if bc_num else "",
            "booking_confirmation_date": parse_date(ws.cell(row, col_bc_date).value if col_bc_date else None),
            "course_code": str(ws.cell(row, col_course).value).strip() if col_course and not is_empty(ws.cell(row, col_course).value) else "",
            "status": str(ws.cell(row, col_status).value).strip() if col_status and not is_empty(ws.cell(row, col_status).value) else "",
            "monthly_fee_after_discount": str(ws.cell(row, col_monthly_fee_after_discount).value).strip() if col_monthly_fee_after_discount and not is_empty(ws.cell(row, col_monthly_fee_after_discount).value) else "",
            "source_file": os.path.basename(path),
        })

    wb.close()
    return records


def load_source_two_file(path, name_to_records):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws, must_contain=["student name"], search_rows=15)
    if header_row is None:
        wb.close()
        return 0

    headers = build_header_map(ws, header_row)

    col_name   = get_col(headers, "Student Name")
    col_birth  = get_col(headers, "Birth date", "Birth Date", "Date of Birth", "DOB")
    col_addr1  = get_col(headers, "Address 1", "Address1", "Address")
    col_addr2  = get_col(headers, "Address 2", "Address2")
    col_mobile = get_col(headers, "Mobile No.", "Mobile Number", "Mobile", "Mobile No")
    col_course = get_col(headers, "Course Enrolled", "Course")
    col_title  = get_col(headers, "Title")
    col_monthly_fee_after_discount = get_col(headers, "Monthly Fees After Discount")
    col_father = get_col(
        headers,
        "Guardian Name", "Father Name", "Fathername",
        "Father's Name", "Parent Name", "Guardian",
    )

    count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        name_raw = ws.cell(row, col_name).value if col_name else None
        if is_empty(name_raw):
            continue

        norm = normalize_name(name_raw)
        if not norm:
            continue

        father_val = ws.cell(row, col_father).value if col_father else None
        if is_empty(father_val) or str(father_val).strip() == "--":
            father_val = ""

        mobile_val = ws.cell(row, col_mobile).value if col_mobile else None
        if is_empty(mobile_val):
            mobile_val = ""

        rec = {
            "student_name_norm": norm,
            "student_name_clean": clean_display_name(name_raw),
            "birth_date": parse_date(ws.cell(row, col_birth).value if col_birth else None),
            "address_1": ws.cell(row, col_addr1).value if col_addr1 else None,
            "address_2": ws.cell(row, col_addr2).value if col_addr2 else None,
            "mobile": str(mobile_val).strip() if mobile_val else "",
            "father_name": str(father_val).strip() if father_val else "",
            "course_enrolled": str(ws.cell(row, col_course).value).strip()
                if col_course and not is_empty(ws.cell(row, col_course).value) else "",
            "title": str(ws.cell(row, col_title).value).strip()
                if col_title and not is_empty(ws.cell(row, col_title).value) else "",
            "monthly_fee_after_discount" :  str(ws.cell(row, col_monthly_fee_after_discount).value).strip()
                if col_monthly_fee_after_discount and not is_empty(ws.cell(row, col_monthly_fee_after_discount).value) else "",
            "source_file": os.path.basename(path),
            
        }
        rec["full_address"] = combine_address(rec["address_1"], rec["address_2"])
        name_to_records[norm].append(rec)
        count += 1

    wb.close()
    return count


def find_best_personal_match(booking_rec, name_to_personal):
    norm = booking_rec["student_name_norm"]
    if not norm:
        return None

    candidates = name_to_personal.get(norm, [])
    if not candidates:
        return None

    course = booking_rec.get("course_code", "")
    if course:
        for c in candidates:
            if course in (c.get("course_enrolled") or ""):
                return c

    return max(
        candidates,
        key=lambda x: (
            bool(x.get("birth_date")),
            bool(x.get("full_address")),
            bool(x.get("mobile")),
            bool(x.get("father_name")),
        ),
    )


# ============================================================
# CORE EXTRACTION
# ============================================================
def run_data_extraction(log_callback=None, progress_callback=None):
    
    def log(msg):
        if log_callback:
            log_callback(msg)

    def progress(pct):
        if progress_callback:
            progress_callback(pct)

    for d in (SOURCE_DIR, TEMPLATE_DIR, DEST_DIR):
        os.makedirs(d, exist_ok=True)

    progress(5)
    log("=" * 60)
    log("  Data Extraction started")
    log("=" * 60)
    log(f"Source folder   : {SOURCE_DIR}")
    log(f"Template folder : {TEMPLATE_DIR}")
    log(f"Destination     : {DEST_DIR}")
    log("")

    # 1. Discover
    source_files = list_xlsx(SOURCE_DIR)
    if not source_files:
        raise RuntimeError("No Excel files found in source/ folder.")

    log(f"[1/6] Found {len(source_files)} Excel file(s) in source/:")
    for p in source_files:
        log(f"       • {os.path.basename(p)}")

    source_one_files, source_two_files = classify_source_files(source_files)
    log(f"\n       Classified as Source File 1  : {len(source_one_files)}")
    for p in source_one_files:
        log(f"         – {os.path.basename(p)}")
    log(f"       Classified as Source File 2 : {len(source_two_files)}")
    for p in source_two_files:
        log(f"         – {os.path.basename(p)}")

    if not source_one_files:
        raise RuntimeError("No Booking Confirmation file detected in source/.")

    progress(15)

    # 2. Load source file 1 data
    log("\n[2/6] Loading Source File 1 data...")
    source_file_records = []
    for path in source_one_files:
        recs = load_source_one_file(path)
        log(f"       {os.path.basename(path)} → {len(recs)} records")
        source_file_records.extend(recs)
    log(f"       Total source file one records : {len(source_file_records)}")
    progress(35)

    # 3. Load source file 2
    log("\n[3/6] Loading Source File 2 data...")
    name_to_personal = defaultdict(list)
    total_personal = 0
    for path in source_two_files:
        cnt = load_source_two_file(path, name_to_personal)
        log(f"       {os.path.basename(path)} → {cnt} records")
        total_personal += cnt
    log(f"       Total source file two records : {total_personal}")
    log(f"       Unique normalized names: {len(name_to_personal)}")
    progress(55)

    # 4. Template
    log("\n[4/6] Locating destination template...")
    template_files = list_xlsx(TEMPLATE_DIR)
    if not template_files:
        raise RuntimeError("No template Excel file found in template/ folder.")
    template_path = template_files[0]
    log(f"       Using template: {os.path.basename(template_path)}")
    progress(60)

    # 5. Match
    log("\n[5/6] Matching and preparing output rows...")
    log("       (Duplicate check DISABLED – every booking row will be written)")

    dest_columns = [
        "Class", "Section", "Studentname", "Fathername", "Dateofbirth",
        "Familynumber", "Admissiondate", "Genderid",
        "Admissionnumber/ Registration No.", "Rollnumber",
        "Guardiancnic", "Mobilenumber", "Studentaddress",
        "Tuition  Fee |1", "Previous Balance",
    ]

    output_rows = []
    matched = 0
    unmatched = 0
    total = len(source_file_records) or 1

    for i, b in enumerate(source_file_records):
        p = find_best_personal_match(b, name_to_personal)
        if p:
            matched += 1
        else:
            unmatched += 1

        genderid = ""
        if p and p.get("title") == "1":
            genderid = "Male"
        elif p and p.get("title") == "2":
            genderid = "Female"

        output_rows.append({
            "Class": "",
            "Section": "",
            "Studentname": b["student_name_clean"],
            "Fathername": p["father_name"] if p else "",
            "Dateofbirth": format_date(p["birth_date"]) if p and p.get("birth_date") else None,
            "Familynumber": b.get("course_code", ""),
            "Admissiondate": format_date(b.get("booking_confirmation_date")),
            "Genderid": genderid,
            "Admissionnumber/ Registration No.": b.get("booking_confirmation_no", ""),
            "Rollnumber": b.get("enrollment_no", ""),
            "Guardiancnic": "",
            "Mobilenumber": p["mobile"] if p else "",
            "Studentaddress": p["full_address"] if p else "",
            "Tuition  Fee |1": b["monthly_fee_after_discount"],
            "Previous Balance": "",
        })

        if i % 200 == 0:
            progress(60 + 25 * (i / total))

    progress(85)
    log(f"       Rows prepared : {len(output_rows)}  (Matched: {matched}, Unmatched: {unmatched})")

    # 6. Write
    log("\n[6/6] Writing result file...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = f"Client_Data_Collection_{timestamp}.xlsx"
    out_path = os.path.join(DEST_DIR, out_name)

    wb = load_workbook(template_path)
    ws = wb.active

    if ws.max_row > 1:
        for r in range(2, ws.max_row + 1):
            for c in range(1, len(dest_columns) + 1):
                ws.cell(r, c).value = None

    for idx, rowdata in enumerate(output_rows):
        excel_row = idx + 2
        for col_idx, col_name in enumerate(dest_columns, start=1):
            value = rowdata.get(col_name)
            ws.cell(excel_row, col_idx).value = value if value not in (None, "") else None

    wb.save(out_path)
    wb.close()
    progress(95)

    # Log file
    with open(LOG_FILE, "w", encoding="utf-8") as f:
        f.write(f"Data Extraction Log – {datetime.now().isoformat()}\n")
        f.write("=" * 60 + "\n")
        f.write(f"Source file One used     : {len(source_one_files)}\n")
        for p in source_one_files:
            f.write(f"  - {os.path.basename(p)}\n")
        f.write(f"Source file Two used    : {len(source_two_files)}\n")
        for p in source_two_files:
            f.write(f"  - {os.path.basename(p)}\n")
        f.write(f"Total source file one records  : {len(source_file_records)}\n")
        f.write(f"Total source file two records : {total_personal}\n")
        f.write(f"Rows written           : {len(output_rows)}\n")
        f.write(f"Matched with personal  : {matched}\n")
        f.write(f"Unmatched (booking only): {unmatched}\n")
        f.write(f"Output file            : {out_name}\n")
        f.write("\nNOTE: Duplicate checking is DISABLED.\n")
        f.write("      Every booking row is written even if names repeat.\n")

    progress(100)
    log("\n" + "=" * 60)
    log("  DATA Extraction COMPLETED SUCCESSFULLY")
    log("=" * 60)
    log(f"  Rows written     : {len(output_rows)}")
    log(f"  Matched          : {matched}")
    log(f"  Unmatched        : {unmatched}")
    log(f"  Output file      : {out_path}")
    log(f"  Log file         : {LOG_FILE}")
    log("=" * 60)
    log("\nTemplate file was NOT modified.")

    return {
        "rows": len(output_rows),
        "matched": matched,
        "unmatched": unmatched,
        "output_path": out_path,
        "output_name": out_name,
        "log_file": LOG_FILE,
    }


# ============================================================
# GUI
# ============================================================
class ConsolidationApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Data Extraction")
        self.geometry("820x680")
        self.minsize(700, 560)
        self.configure(bg="#f0f2f5")

        # Style
        style = ttk.Style(self)
        if "vista" in style.theme_names():
            style.theme_use("vista")
        elif "clam" in style.theme_names():
            style.theme_use("clam")

        style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"), background="#f0f2f5")
        style.configure("Header.TLabel", font=("Segoe UI", 10, "bold"), background="#f0f2f5")
        style.configure("Info.TLabel", font=("Segoe UI", 9), background="#f0f2f5", foreground="#444")
        style.configure("Run.TButton", font=("Segoe UI", 11, "bold"), padding=8)
        style.configure("TButton", font=("Segoe UI", 9), padding=4)
        style.configure("TLabelframe", background="#f0f2f5")
        style.configure("TLabelframe.Label", font=("Segoe UI", 9, "bold"), background="#f0f2f5")
        style.configure("TFrame", background="#f0f2f5")
        style.configure("Status.TLabel", font=("Segoe UI", 9), background="#e8eaed", foreground="#333")

        self._running = False
        self._build_ui()
        self._refresh_file_lists()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_ui(self):
        # ---- Header ----
        hdr = ttk.Frame(self, padding=(16, 12, 16, 4))
        hdr.pack(fill=tk.X)
        ttk.Label(hdr, text="Data Extraction", style="Title.TLabel").pack(side=tk.LEFT)
        ##ttk.Label(hdr, text="Booking + Personal → Client Data Collection", style="Info.TLabel").pack(side=tk.LEFT, padx=(12, 0))

        # ---- Folders section ----
        folders = ttk.LabelFrame(self, text="  Folders  ", padding=10)
        folders.pack(fill=tk.X, padx=16, pady=(8, 4))

        # Source
        row1 = ttk.Frame(folders)
        row1.pack(fill=tk.X, pady=2)
        ttk.Label(row1, text="Source:", width=10, style="Header.TLabel").pack(side=tk.LEFT)
        self.src_var = tk.StringVar(value=SOURCE_DIR)
        ttk.Entry(row1, textvariable=self.src_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        ttk.Button(row1, text="Open", width=8, command=lambda: self._open_folder(SOURCE_DIR)).pack(side=tk.LEFT, padx=2)
        ttk.Button(row1, text="Refresh", width=8, command=self._refresh_file_lists).pack(side=tk.LEFT)

        # Template
        row2 = ttk.Frame(folders)
        row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="Template:", width=10, style="Header.TLabel").pack(side=tk.LEFT)
        self.tpl_var = tk.StringVar(value=TEMPLATE_DIR)
        ttk.Entry(row2, textvariable=self.tpl_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        ttk.Button(row2, text="Open", width=8, command=lambda: self._open_folder(TEMPLATE_DIR)).pack(side=tk.LEFT, padx=2)

        # Destination
        row3 = ttk.Frame(folders)
        row3.pack(fill=tk.X, pady=2)
        ttk.Label(row3, text="Destination:", width=10, style="Header.TLabel").pack(side=tk.LEFT)
        self.dst_var = tk.StringVar(value=DEST_DIR)
        ttk.Entry(row3, textvariable=self.dst_var, state="readonly").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        ttk.Button(row3, text="Open", width=8, command=lambda: self._open_folder(DEST_DIR)).pack(side=tk.LEFT, padx=2)

        # ---- Detected files ----
        files_frame = ttk.Frame(self)
        files_frame.pack(fill=tk.X, padx=16, pady=4)

        left = ttk.LabelFrame(files_frame, text="  Source File One Detected  ", padding=6)
        left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 4))
        self.source_file_one_list = tk.Listbox(left, height=4, font=("Verdana", 9), activestyle="none",
                                       selectmode=tk.EXTENDED, bg="#ffffff", relief=tk.FLAT,
                                       highlightthickness=1, highlightbackground="#ccc")
        self.source_file_one_list.pack(fill=tk.BOTH, expand=True)

        right = ttk.LabelFrame(files_frame, text="  Source File Two Detected  ", padding=6)
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 0))
        self.source_file_two_list = tk.Listbox(right, height=4, font=("Verdana", 9), activestyle="none",
                                        selectmode=tk.EXTENDED, bg="#ffffff", relief=tk.FLAT,
                                        highlightthickness=1, highlightbackground="#ccc")
        self.source_file_two_list.pack(fill=tk.BOTH, expand=True)

        # Template label
        self.tpl_label = ttk.Label(self, text="Template: (none)", style="Info.TLabel")
        self.tpl_label.pack(anchor=tk.W, padx=20, pady=(0, 4))

        # ---- Run + Progress ----
        action = ttk.Frame(self, padding=(16, 4))
        action.pack(fill=tk.X)
        self.run_btn = ttk.Button(action, text="▶  Run Data Extraction", style="Run.TButton",
                                  command=self._start_run)
        self.run_btn.pack(side=tk.LEFT)
        self.open_out_btn = ttk.Button(action, text="Open Output Folder", command=self._open_output_folder)
        self.open_out_btn.pack(side=tk.LEFT, padx=8)
        self.open_log_btn = ttk.Button(action, text="View Log", command=self._open_log)
        self.open_log_btn.pack(side=tk.LEFT)

        self.progress = ttk.Progressbar(self, mode="determinate", maximum=100)
        self.progress.pack(fill=tk.X, padx=16, pady=(6, 2))

        # ---- Log area ----
        log_frame = ttk.LabelFrame(self, text="  Process Log  ", padding=6)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=(4, 8))
        self.log_text = scrolledtext.ScrolledText(
            log_frame, height=14, font=("Verdana", 9),
            bg="#1e1e1e", fg="#d4d4d4", insertbackground="#fff",
            relief=tk.FLAT, wrap=tk.WORD, state=tk.DISABLED
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        self.log_text.tag_configure("success", foreground="#4ec9b0")
        self.log_text.tag_configure("error", foreground="#f44747")
        self.log_text.tag_configure("info", foreground="#9cdcfe")

        # ---- Status bar ----
        self.status_var = tk.StringVar(value="Ready – place Excel files in source/ and template/, then click Run")
        status = ttk.Label(self, textvariable=self.status_var, style="Status.TLabel", padding=(10, 4), relief=tk.SUNKEN)
        status.pack(fill=tk.X, side=tk.BOTTOM)

        self._last_output = None

    # ---- Helpers ----
    def _log(self, msg, tag=None):
        self.log_text.configure(state=tk.NORMAL)
        if tag:
            self.log_text.insert(tk.END, msg + "\n", tag)
        else:
            self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _set_progress(self, pct):
        self.progress["value"] = pct
        self.update_idletasks()

    def _open_folder(self, path):
        os.makedirs(path, exist_ok=True)
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')

    def _refresh_file_lists(self):
        self.source_file_one_list.delete(0, tk.END)
        self.source_file_two_list.delete(0, tk.END)

        source_files = list_xlsx(SOURCE_DIR)
        if source_files:
            booking, personal = classify_source_files(source_files)
            for p in booking:
                self.source_file_one_list.insert(tk.END, os.path.basename(p))
            for p in personal:
                self.source_file_two_list.insert(tk.END, os.path.basename(p))
        else:
            self.source_file_one_list.insert(tk.END, "(no files – put .xlsx files in source/)")
            self.source_file_two_list.insert(tk.END, "(no files – put .xlsx files in source/)")

        templates = list_xlsx(TEMPLATE_DIR)
        if templates:
            self.tpl_label.configure(text=f"Template:  {os.path.basename(templates[0])}")
        else:
            self.tpl_label.configure(text="Template:  (none found in template/ folder)")

        self.status_var.set(f"Ready – {len(source_files)} source file(s), {len(templates)} template(s)")

    def _start_run(self):
        if self._running:
            return
        self._running = True
        self.run_btn.configure(state=tk.DISABLED)
        self.progress["value"] = 0
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.status_var.set("Running Data Extraction… please wait")

        def worker():
            try:
                result = run_data_extraction(
                    log_callback=lambda m: self.after(0, self._log, m),
                    progress_callback=lambda p: self.after(0, self._set_progress, p),
                )
                self.after(0, self._on_success, result)
            except Exception as e:
                tb = traceback.format_exc()
                self.after(0, self._on_error, str(e), tb)

        threading.Thread(target=worker, daemon=True).start()

    def _on_success(self, result):
        self._running = False
        self.run_btn.configure(state=tk.NORMAL)
        self._last_output = result["output_path"]
        self.status_var.set(
            f"Done – {result['rows']} rows written  |  Matched: {result['matched']}  |  Unmatched: {result['unmatched']}"
        )
        self._log("\n✓  Data Extraction finished successfully.", "success")
        messagebox.showinfo(
            "Success",
            f"Data Extraction completed!\n\n"
            f"Rows written : {result['rows']}\n"
            f"Matched      : {result['matched']}\n"
            f"Unmatched    : {result['unmatched']}\n\n"
            f"Output:\n{result['output_name']}"
        )

    def _on_error(self, msg, tb):
        self._running = False
        self.run_btn.configure(state=tk.NORMAL)
        self.progress["value"] = 0
        self.status_var.set("Error – see log for details")
        self._log(f"\n✗  ERROR: {msg}", "error")
        self._log(tb, "error")
        messagebox.showerror("Error", f"Data Extraction failed:\n\n{msg}")

    def _open_output_folder(self):
        """Open the destination folder so the user can see all generated output files."""
        self._open_folder(DEST_DIR)

    def _open_log(self):
        if os.path.isfile(LOG_FILE):
            if sys.platform == "win32":
                os.startfile(LOG_FILE)
            else:
                # show in log area
                with open(LOG_FILE, encoding="utf-8") as f:
                    content = f.read()
                self.log_text.configure(state=tk.NORMAL)
                self.log_text.delete(1.0, tk.END)
                self.log_text.insert(tk.END, content)
                self.log_text.configure(state=tk.DISABLED)
        else:
            messagebox.showinfo("Log", "No log file yet. Run Data Extraction first.")

    def _on_close(self):
        if self._running:
            if not messagebox.askyesno("Busy", "Data Extraction is still running. Exit anyway?"):
                return
        self.destroy()


# ============================================================
# ENTRY POINT
# ============================================================
def main():
    # Ensure folders exist
    for d in (SOURCE_DIR, TEMPLATE_DIR, DEST_DIR):
        os.makedirs(d, exist_ok=True)

    # If launched with --cli flag, keep original console behaviour
    if "--cli" in sys.argv:
        def console_log(msg):
            print(msg)
        try:
            run_data_extraction(log_callback=console_log)
        except Exception as e:
            print(f"\nERROR: {e}")
            input("\nPress Enter to exit...")
        return

    app = ConsolidationApp()
    app.mainloop()


if __name__ == "__main__":
    main()
